from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def create_xlsx(expenses):
    df = pd.DataFrame(expenses)
    
    columns_order = [
        "id",
        "name",
        "amount_uah",
        "amount_usd",
        "date"
    ]
    
    df = df[columns_order]
    
    df.rename(columns={
        "id": "ID",
        "name": "Назва витрати",
        "amount_uah": "Сума витрати (UAH)",
        "amount_usd": "Сума витрати (USD)",
        "date": "Дата"
    }, inplace=True)
    
    df['Дата'] = pd.to_datetime(df['Дата']).dt.strftime('%d.%m.%Y')
    
    with pd.ExcelWriter("expenses.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Expenses")
        
        workbook = writer.book
        sheet = workbook['Expenses']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '#,##0.00'
    
    return "expenses.xlsx"
